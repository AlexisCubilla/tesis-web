"""API del taller de experimentación y servidor de la web didáctica.

El backend es una capa delgada: recibe una configuración, arma los objetos de `tesis`, llama a la
etapa que corresponde y guarda el resultado con su hash. **Toda la lógica científica vive en el
paquete de la tesis** (ver `etapas.py`).
"""

from __future__ import annotations

import dataclasses
import json
import shutil
import tempfile
from pathlib import Path

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from starlette.background import BackgroundTask

from . import ajustes, almacen, etapas, proxy_oo, revision, trabajos

RAIZ = ajustes.RAIZ
WEB = ajustes.DIR_WEB

app = FastAPI(title="Taller — Etapa 1", version="0.1.0")
_con = almacen.conectar()
revision.preparar()
trabajos.iniciar_trabajador(_con)


class PeticionEtiqueta(BaseModel):
    etiqueta: str | None = None


class PeticionEjecutar(BaseModel):
    etapa: str
    padre: str | None = None
    parametros: dict = {}
    etiqueta: str | None = None


# --------------------------------------------------------------------------------------
# API
# --------------------------------------------------------------------------------------

# Compresión. `StaticFiles` no comprime nada por su cuenta, así que hasta acá
# `echarts.min.js` viajaba en 1,03 MB por la red: es el archivo más grande del
# sitio por un orden de magnitud y es texto, o sea que comprime muy bien. El
# umbral de 900 bytes es el de la biblioteca: por debajo de eso el encabezado y el
# trabajo de comprimir cuestan más que lo que se ahorra.
app.add_middleware(GZipMiddleware, minimum_size=900)


@app.get("/api/etapas")
def get_etapas():
    """Las etapas con sus parámetros. El frontend arma los formularios a partir de esto."""
    return {"cadena": list(etapas.CADENA), "etapas": etapas.descripcion_serializable()}


@app.get("/api/estado")
def get_estado():
    """Estado general: si están los datos base, con qué versión de la tesis corre y su configuración."""
    try:
        import tesis

        version = tesis.__version__
    except Exception:
        version = None
    return {
        "datos_listos": almacen.TABLA_CRUDA.exists(),
        "commit_tesis": almacen.commit_tesis(),
        "version_tesis": version,
        "ruta_tesis": str(ajustes.ruta_tesis()),
        "nodos": len(almacen.listar(_con)),
        # La configuración de referencia se sirve desde el backend para que exista en un solo lugar
        # (`backend/ajustes.py`, ajustable por .env) y no duplicada en el frontend.
        "config_tesis": ajustes.CONFIG_TESIS,
        "origen_config": ajustes.origen_config(),
        "muestreo_segundos": ajustes.MUESTREO_SEGUNDOS,
    }


@app.get("/api/arbol")
def get_arbol():
    """Todos los nodos del árbol de ejecuciones."""
    return {"nodos": almacen.listar(_con)}


@app.post("/api/ejecutar")
def post_ejecutar(pet: PeticionEjecutar):
    """Crea (o recupera) un nodo y lo encola.

    Si la configuración ya se ejecutó antes, devuelve el nodo existente sin recalcular: es el caché
    por hash. Ramificar es simplemente pedir una etapa con otros parámetros sobre el mismo padre.
    """
    if pet.etapa not in etapas.REGISTRO:
        raise HTTPException(404, f"Etapa desconocida: {pet.etapa}")

    etapa = etapas.REGISTRO[pet.etapa]
    params = {**etapa.defectos(), **(pet.parametros or {})}
    params = {k: (tuple(v) if isinstance(v, list) else v) for k, v in params.items()}

    # Validación de la cadena: la etapa pedida debe seguir a la del padre.
    padre = almacen.obtener(_con, pet.padre) if pet.padre else None
    if pet.padre and padre is None:
        raise HTTPException(404, f"Nodo padre inexistente: {pet.padre}")
    esperada = etapas.etapa_siguiente(padre["etapa"] if padre else None)
    if pet.etapa != esperada:
        raise HTTPException(400, f"Después de {padre['etapa'] if padre else 'la raíz'} "
                                 f"corresponde '{esperada}', no '{pet.etapa}'")

    clave = almacen.clave_nodo(pet.etapa, params, pet.padre)
    existente = almacen.obtener(_con, clave)
    if existente and existente["estado"] == "listo" and existente["tiene_resultado"]:
        return {"clave": clave, "reutilizado": True, "nodo": existente}

    almacen.crear_nodo(_con, clave, pet.etapa, pet.padre, params, pet.etiqueta)
    trabajos.encolar(clave)
    return {"clave": clave, "reutilizado": False, "nodo": almacen.obtener(_con, clave)}


@app.get("/api/nodo/{clave}")
def get_nodo(clave: str):
    nodo = almacen.obtener(_con, clave)
    if nodo is None:
        raise HTTPException(404, "Nodo inexistente")
    nodo["cadena"] = almacen.cadena_hasta(_con, clave)
    return nodo


@app.delete("/api/nodo/{clave}")
def delete_nodo(clave: str, forzar: bool = False):
    """Borra el nodo, todo lo que cuelga de él y sus resultados en disco.

    **Se planta si el experto ya comentó alguna de esas ramas.** Todo lo demás que borra este
    endpoint se puede volver a calcular corriendo el pipeline otra vez; lo que escribió el experto,
    no. Y el botón de borrar está en la misma pantalla donde se comenta, así que un clic en el nodo
    equivocado se llevaría trabajo irrepetible sin avisar.

    Con `forzar=1` se borra igual, pero el documento no se va con la rama: queda huérfano y se puede
    seguir descargando (ver `revision.desvincular`).
    """
    if almacen.obtener(_con, clave) is None:
        raise HTTPException(404, "Nodo inexistente")

    # El borrado es en cascada, así que hay que mirar el nodo y todo lo que cuelga de él.
    claves = [clave, *almacen.descendientes(_con, clave)]
    comentados = revision.con_trabajo_del_experto(claves)

    if comentados and not forzar:
        raise HTTPException(409, {
            "mensaje": "Esa rama tiene documentos con comentarios del experto. Eso no se puede "
                       "recalcular: confirmá el borrado si de verdad querés seguir.",
            "documentos": [
                {"id": d["id"], "nodo": d["nodo"], "tipo": d["tipo"], "nombre": d["nombre"],
                 "version": d["version"], "guardado_en": d["guardado_en"],
                 "ultimo_autor": d["ultimo_autor"]}
                for d in comentados
            ],
        })

    huerfanos = revision.desvincular(claves) if comentados else 0
    resultado = almacen.borrar(_con, clave)
    resultado["documentos_huerfanos"] = huerfanos
    return resultado


#: Tope de puntos de una serie temporal antes de empezar a submuestrear. Ver `get_datos`.
SERIE_COMPLETA = 20_000


@app.put("/api/nodo/{clave}/etiqueta")
def put_etiqueta(clave: str, pet: PeticionEtiqueta):
    """Renombra un paso. No toca el hash: es un nombre para humanos, no parte de su identidad."""
    if almacen.obtener(_con, clave) is None:
        raise HTTPException(404, "Nodo inexistente")
    nombre = (pet.etiqueta or "").strip()[:60] or None
    almacen.etiquetar(_con, clave, nombre)
    return {"clave": clave, "etiqueta": nombre}


@app.get("/api/nodo/{clave}/datos")
def get_datos(clave: str, hoja: str | None = None, limite: int = 300):
    """Datos para inspeccionar una etapa. Siempre acotados: esto viaja al navegador."""
    import numpy as np

    nodo = almacen.obtener(_con, clave)
    if nodo is None:
        raise HTTPException(404, "Nodo inexistente")
    if nodo["estado"] != "listo" or not nodo["tiene_resultado"]:
        raise HTTPException(409, "El nodo todavía no tiene resultado")

    salida = almacen.cargar_resultado(clave)
    etapa = nodo["etapa"]
    grupo = "SheetName"

    # ---- Datos: la serie temporal real de una hoja -------------------------------------
    if etapa == "datos":
        hojas = salida[grupo].drop_duplicates().tolist()
        elegida = hoja if hoja in hojas else hojas[0]
        sub = salida[salida[grupo] == elegida]
        señales = [c for c in sub.columns if c not in (grupo, "segment")]
        # El gráfico deja acercarse a un tramo, y acercarse sobre una curva diezmada solo agranda
        # la diezma: no aparece ni un dato nuevo. Por eso la serie va entera. La hoja más grande
        # tiene ~1.100 mediciones × 3 señales, unas decenas de KB contra un servidor local; el
        # submuestreo queda solo como red de contención por si alguna vez hay una hoja enorme.
        paso = max(1, len(sub) // max(limite, SERIE_COMPLETA))
        sub = sub.iloc[::paso]
        return {
            "tipo": "serie",
            "hojas": hojas,
            "hoja": elegida,
            "senales": señales,
            "valores": [[None if pd_isna(v) else float(v) for v in fila]
                        for fila in sub[señales].to_numpy()],
            "por_hoja": salida.groupby(grupo).size().sort_values(ascending=False).head(20).to_dict(),
        }

    # ---- Ventaneo: ventanas de ejemplo dibujables --------------------------------------
    if etapa == "ventaneo":
        ventanas, meta = salida["ventanas"], salida["meta"]
        n = len(ventanas)
        indices = [0, n // 4, n // 2, (3 * n) // 4][:max(1, min(4, n))] if n else []
        ejemplos = []
        for i in indices:
            ejemplos.append({
                "pos": int(i),
                "hoja": str(meta.loc[i, grupo]),
                "inicio": int(meta.loc[i, "start"]),
                "valores": ventanas.values[i].tolist(),
            })
        return {
            "tipo": "ventanas",
            "senales": list(ventanas.signal_columns),
            "ejemplos": ejemplos,
            "por_hoja": meta.groupby(grupo).size().sort_values(ascending=False).head(20).to_dict(),
            "tamano": int(ventanas.values.shape[1]) if n else 0,
        }

    # ---- Características: qué se calculó, por familia ----------------------------------
    if etapa == "features":
        tabla = salida["features"]
        cols = [c for c in tabla.columns if c != grupo]
        por_senal: dict[str, list[str]] = {}
        for c in cols:
            señal = c.split("__")[0] if "__" in c else "otras"
            por_senal.setdefault(señal, []).append(c.split("__")[-1])
        return {
            "tipo": "features",
            "total": len(cols),
            "por_senal": por_senal,
            # La forma de cada distribución, en desvíos estándar. Estandarizar no es maquillaje:
            # es lo que hace que el eje tenga UN significado para las 72 —«a cuántos desvíos de su
            # media»— en vez de mezclar volts, miliamperios y grados en la misma regla.
            "cajas": _cajas_estandarizadas(tabla[sorted(cols)]),
        }

    # ---- Filtrado: cuáles sobrevivieron, cuáles no y por qué ---------------------------
    if etapa == "filtrado":
        previas = salida["features"]
        cols = sorted(c for c in previas.columns if c != grupo)
        despues = {c for c in salida["filtradas"].columns if c != grupo}

        from tesis import filtering
        from tesis.config import CONFIG

        # La MISMA función que usó la ejecución. Antes esto armaba la config por su
        # cuenta, duplicando la lógica: bastaba con que una de las dos cambiara para
        # que el gráfico de motivos siguiera pareciendo correcto y dijera que se
        # descartó algo que no se descartó. Ver la nota de `etapas.config_filtrado`.
        cfg = etapas.config_filtrado(nodo["parametros"])
        # El motivo del descarte se saca corriendo el primer filtro por separado, con las mismas
        # funciones del paquete de la tesis. Importa distinguirlos: los dos filtros van en cadena,
        # así que lo que se cae por IQR ni siquiera llega a evaluarse por correlación.
        tras_iqr = {c for c in filtering.filter_low_variance(previas, cfg).columns if c != grupo}
        motivo = {c: ("conservada" if c in despues
                      else "poca variación" if c not in tras_iqr
                      else "repite a otra") for c in cols}

        # ⚠ ACOPLAMIENTO CONOCIDO CON `filtering.filter_low_variance`.
        #
        # El MOTIVO de cada descarte se delega (arriba), pero el NÚMERO que se dibuja se calcula
        # acá: el rango intercuartílico y el corte por percentil, replicando el criterio que hoy usa
        # esa función. Coinciden — se verificó leyéndola antes de graficar, y por eso el gráfico
        # dice «rango intercuartílico» y no «varianza».
        #
        # Pero si en la tesis se cambia ese criterio (a desviación estándar, a MAD, a lo que sea),
        # las etiquetas de motivo seguirían correctas y estas barras pasarían a mostrar una magnitud
        # que el filtro ya no usa: un gráfico que sigue pareciendo correcto mientras miente. Es el
        # modo de falla que el ADR A1 existe para evitar.
        #
        # QUIEN TOQUE `filter_low_variance` TIENE QUE TOCAR ESTO TAMBIÉN. La salida de fondo es que
        # el paquete exponga qué medida y qué corte usó, en vez de que el taller lo deduzca; ese
        # cambio va en el repo de la tesis con su ADR (ver A1).
        X = previas[cols]
        iqr = (X.quantile(0.75) - X.quantile(0.25)).astype(float)
        corte_pct = (float(np.nanpercentile(iqr.to_numpy(), cfg.lowvar_percentile))
                     if cfg.lowvar_percentile and cfg.lowvar_percentile > 0 else None)
        corr = X.corr(method=cfg.correlation_method).fillna(0.0)

        return {
            "tipo": "filtrado",
            "conservadas": sorted(despues),
            "descartadas": sorted(set(cols) - despues),
            "features": cols,
            "motivo": [motivo[c] for c in cols],
            # El filtro mira el rango intercuartílico, no la varianza. Se manda el mismo número que
            # mira, para que el corte dibujado sea el corte real.
            "iqr": [round(float(iqr[c]), 6) for c in cols],
            "corte_iqr_min": float(cfg.iqr_min),
            "corte_percentil": corte_pct,
            "percentil": float(cfg.lowvar_percentile),
            "umbral_correlacion": float(cfg.correlation_threshold),
            "correlacion": [[round(float(v), 4) for v in fila] for fila in corr.to_numpy()],
        }

    # ---- Detección: distribución de puntajes y las ventanas más extremas ---------------
    if etapa == "deteccion":
        sc = salida["scores"]
        meta = salida["meta"]
        top = {}
        for c in sc.columns:
            idx = sc[c].nlargest(8).index
            top[c] = [{"pos": int(p), "hoja": str(meta.loc[p, grupo]),
                       "inicio": int(meta.loc[p, "start"]), "score": float(sc.loc[p, c])}
                      for p in idx]
        # Sin percentiles: eran la misma información que los histogramas de la pestaña de análisis
        # —una curva acumulada y una densidad dicen lo mismo—, y allá además está el corte de
        # candidatos marcado. Dos vistas del mismo dato hacen que ninguna se mire.
        return {"tipo": "scores", "detectores": list(sc.columns), "top": top}

    # ---- Eventos: el entregable --------------------------------------------------------
    if etapa == "eventos":
        tabla = salida["eventos"]
        meta = salida.get("meta")
        # Para la línea de tiempo hace falta el largo del registro de cada hoja, que no está en la
        # tabla de eventos: sale de dónde empieza la última ventana más el tamaño de la ventana.
        pistas = []
        if meta is not None and grupo in meta:
            largo_ventana = int(salida.get("tamano_ventana") or 0)
            for hoja, sub in meta.groupby(grupo):
                pistas.append({"hoja": str(hoja), "desde": int(sub["start"].min()),
                               "hasta": int(sub["start"].max()) + largo_ventana})
            pistas.sort(key=lambda p: p["hasta"] - p["desde"], reverse=True)
        return {
            "tipo": "eventos",
            "columnas": list(tabla.columns),
            "filas": json.loads(tabla.head(limite).to_json(orient="records")),
            "pistas": pistas,
        }

    return {"tipo": "desconocido"}


def pd_isna(v) -> bool:
    import pandas as pd

    return bool(pd.isna(v))


def _cajas_estandarizadas(X) -> list[dict]:
    """Caja y bigotes de cada columna, en desvíos estándar respecto de su propia media.

    Sin estandarizar no se pueden poner juntas: una característica está en volts, otra en
    miliamperios y otra es una energía espectral de seis cifras, así que el eje no querría decir
    nada. Estandarizadas, el eje significa una sola cosa para todas —a cuántos desvíos de su media
    cae el valor— y lo que se compara es la forma: cuál tiene cola larga y cuál no.
    """
    import numpy as np

    salida = []
    for col in X.columns:
        v = X[col].to_numpy(dtype=float)
        v = v[np.isfinite(v)]
        if v.size == 0:
            continue
        sd = float(v.std()) or 1.0
        z = (v - float(v.mean())) / sd
        q1, med, q3 = (float(np.quantile(z, q)) for q in (0.25, 0.5, 0.75))
        rango = q3 - q1
        bajo, alto = q1 - 1.5 * rango, q3 + 1.5 * rango
        dentro = z[(z >= bajo) & (z <= alto)]
        salida.append({
            "nombre": col,
            "caja": [round(float(dentro.min()) if dentro.size else float(z.min()), 3),
                     round(q1, 3), round(med, 3), round(q3, 3),
                     round(float(dentro.max()) if dentro.size else float(z.max()), 3)],
            "atipicos": int(((z < bajo) | (z > alto)).sum()),
        })
    return salida


@app.get("/api/nodo/{clave}/analisis")
def get_analisis(clave: str, detector: str | None = None):
    """Datos para la pantalla de análisis: los cuatro cortes que el taller no mostraba.

    REGLA. Nada de esto redefine el pipeline (ADR A1). La selección de candidatos se le pide al
    propio paquete de la tesis —`export._detector_candidates`, la misma que usa
    `build_candidate_table`— justamente para no tener dos reglas de «qué tramo está marcado» que
    puedan separarse con el tiempo. Lo que se calcula acá es solo cómo mostrarlo: contar
    intersecciones, armar histogramas, correlacionar columnas y proyectar a dos dimensiones.

    La proyección 2D **es cálculo nuevo y es solo para mirar**: no alimenta ninguna etapa, no se
    guarda y ningún número de la tesis depende de ella. Se deja dicho para que no se confunda con
    el PCA que sí es un detector del pipeline.
    """
    import numpy as np
    import pandas as pd

    nodo = almacen.obtener(_con, clave)
    if nodo is None:
        raise HTTPException(404, "Nodo inexistente")
    if nodo["estado"] != "listo" or not nodo["tiene_resultado"]:
        raise HTTPException(409, "El nodo todavía no tiene resultado")

    salida = almacen.cargar_resultado(clave)
    if not isinstance(salida, dict) or "scores" not in salida:
        raise HTTPException(
            409, "El análisis necesita un paso de detección o posterior: ahí aparecen los puntajes"
        )

    scores: "pd.DataFrame" = salida["scores"]
    filtradas: "pd.DataFrame" = salida["filtradas"]
    detectores = list(scores.columns)
    n = int(len(scores))
    grupo = "SheetName"

    # ---- Coincidencia entre detectores -------------------------------------------------
    # `fraccion_candidatos` es un parámetro de la etapa de eventos; si se está mirando un nodo de
    # detección todavía no existe, así que se usa el valor de referencia de la tesis.
    fraccion = float(nodo["parametros"].get("fraccion_candidatos")
                     or ajustes.CONFIG_TESIS.get("eventos", {}).get("fraccion_candidatos", 0.01))

    from tesis import export
    from tesis.config import CONFIG

    cfg_det = dataclasses.replace(CONFIG.detection, detectors=tuple(detectores),
                                  candidate_fraction=fraccion)
    marcados = export._detector_candidates(scores, cfg_det)  # dict[detector] -> set de posiciones

    matriz = [[len(marcados[a] & marcados[b]) for b in detectores] for a in detectores]
    cuantos = pd.Series(0, index=scores.index, dtype=int)
    for d in detectores:
        cuantos.loc[list(marcados[d])] += 1
    reparto = {int(k): int(v) for k, v in cuantos[cuantos > 0].value_counts().sort_index().items()}

    # ---- Distribución de puntajes: histograma y caja ------------------------------------
    puntajes = {}
    for d in detectores:
        s = scores[d].to_numpy(dtype=float)
        s = s[np.isfinite(s)]
        cuentas, bordes = np.histogram(s, bins=40)
        q1, med, q3 = (float(np.quantile(s, q)) for q in (0.25, 0.5, 0.75))
        iqr = q3 - q1
        bajo, alto = q1 - 1.5 * iqr, q3 + 1.5 * iqr
        dentro = s[(s >= bajo) & (s <= alto)]
        puntajes[d] = {
            "bordes": [float(x) for x in bordes],
            "cuentas": [int(x) for x in cuentas],
            "caja": [float(dentro.min()) if dentro.size else float(s.min()), q1, med, q3,
                     float(dentro.max()) if dentro.size else float(s.max())],
            "atipicos": int(((s < bajo) | (s > alto)).sum()),
            "corte": float(np.quantile(s, 1 - fraccion)),
        }

    # ---- Correlación entre características ----------------------------------------------
    # Se correlacionan las de ANTES del filtrado, marcando cuáles sobrevivieron: así el mapa
    # muestra *por qué* se descartó cada una, que es lo que la lista de pastillas no dice.
    previas: "pd.DataFrame" = salida.get("features", filtradas)
    cols_prev = [c for c in previas.columns if c != grupo]
    cols_prev.sort()  # el nombre es «señal__medida», así que ordenar agrupa por señal
    corr = previas[cols_prev].corr().fillna(0.0)
    conservadas = {c for c in filtradas.columns if c != grupo}

    # ---- Dispersión 2D de las ventanas ---------------------------------------------------
    elegido = detector if detector in detectores else detectores[0]
    cols_filt = [c for c in filtradas.columns if c != grupo]
    X = filtradas[cols_filt].to_numpy(dtype=float)
    X = np.nan_to_num(X, nan=0.0, posinf=0.0, neginf=0.0)
    centro = X.mean(axis=0)
    escala = X.std(axis=0)
    escala[escala == 0] = 1.0
    Z = (X - centro) / escala
    # SVD en vez de sklearn: son dos componentes y evita arrastrar una dependencia más para dibujar.
    U, S, _ = np.linalg.svd(Z, full_matrices=False)
    proy = U[:, :2] * S[:2]
    varianza = (S ** 2) / float((S ** 2).sum() or 1.0)

    # ---- Qué corre a una candidata del resto -------------------------------------------
    # Se le pide al paquete de la tesis: `select_candidates` para elegir el grupo y `feature_shift`
    # para medir el desplazamiento. Es la misma pareja de funciones que usa
    # `detector_feature_shift_tables`, la que produce el PNG del pipeline. Alcanzaba con parsear la
    # columna `features_top` de la tabla de eventos, pero esa cadena la arma `export` para
    # mostrarla, no para consumirla: atarse a su formato se rompe en silencio el día que cambie.
    from tesis import detection, interpretation

    pos = detection.select_candidates(scores[elegido], cfg_det)
    corrimiento = interpretation.feature_shift(filtradas, pos)
    top = corrimiento.head(22)
    desplazamiento = {
        "features": [str(i) for i in top.index],
        "shift": [round(float(v), 4) for v in top["standardized_mean_shift"]],
        "familia": [str(v) for v in top["feature_family"]],
        "n_candidatas": int(len(pos)),
    }

    # ---- Solo donde hay eventos: ¿el acuerdo es acuerdo, o es duración? -----------------
    # El propio taller advierte que «un evento largo tiene más chances de juntar métodos
    # distintos». Eso plantea la duda y no muestra la evidencia; acá se muestra.
    eventos = None
    if "eventos" in salida and len(salida["eventos"]):
        t = salida["eventos"]
        eventos = {
            "puntos": [{"id": int(r.event_id), "hoja": str(r.SheetName),
                        "ventanas": int(r.n_ventanas), "detectores": int(r.n_detectores)}
                       for r in t.itertuples()],
            "por_acuerdo": {int(k): {"n": int(len(g)), "mediana": float(g["n_ventanas"].median()),
                                     "media": round(float(g["n_ventanas"].mean()), 2)}
                            for k, g in t.groupby("n_detectores")},
        }

    marcado_elegido = marcados[elegido]
    meta = salida.get("meta")
    hojas = (meta[grupo].astype(str).tolist() if meta is not None and grupo in meta
             else [""] * n)

    return {
        "detectores": detectores,
        "detector": elegido,
        "n_ventanas": n,
        "fraccion_candidatos": fraccion,
        "top_k": int(np.ceil(n * fraccion)),
        "coincidencia": {"orden": detectores, "matriz": matriz, "reparto": reparto},
        "puntajes": puntajes,
        "correlacion": {
            "features": cols_prev,
            "matriz": [[round(float(v), 4) for v in fila] for fila in corr.to_numpy()],
            "conservadas": [c in conservadas for c in cols_prev],
        },
        "desplazamiento": desplazamiento,
        "eventos": eventos,
        "dispersion": {
            "x": [round(float(v), 4) for v in proy[:, 0]],
            "y": [round(float(v), 4) for v in proy[:, 1]],
            "puntaje": [round(float(v), 6) for v in scores[elegido].to_numpy(dtype=float)],
            "candidato": [bool(i in marcado_elegido) for i in range(n)],
            "hoja": hojas,
            "varianza_explicada": [round(float(varianza[0]), 4), round(float(varianza[1]), 4)],
        },
    }


@app.get("/api/nodo/{clave}/evento/{event_id}")
def get_evento(clave: str, event_id: int):
    """Las series crudas del tramo de un evento, para dibujarlo."""
    nodo = almacen.obtener(_con, clave)
    if nodo is None or nodo["etapa"] != "eventos" or not nodo["tiene_resultado"]:
        raise HTTPException(404, "No hay eventos en ese nodo")

    salida = almacen.cargar_resultado(clave)
    tabla = salida["eventos"]
    fila = tabla[tabla["event_id"] == event_id]
    if fila.empty:
        raise HTTPException(404, "Evento inexistente")
    fila = fila.iloc[0]

    ventanas = salida["ventanas"]
    meta = salida["meta"]
    posiciones = meta.index[(meta["segment"] == fila["segment"])
                            & (meta["start"] >= fila["start"])
                            & (meta["start"] < fila["end"])].tolist()
    if not posiciones:
        raise HTTPException(404, "Sin ventanas para el evento")

    import numpy as np

    largo = int(fila["end"] - fila["start"])
    señales = list(ventanas.signal_columns)
    acum = np.full((largo, len(señales)), np.nan)
    for p in posiciones:
        desde = int(meta.loc[p, "start"] - fila["start"])
        bloque = ventanas.values[p]
        # Una ventana que arranca cerca del final se extiende más allá del tramo del evento
        # (sobre todo con el límite de ventanas por evento activo): se recorta al rango.
        hasta = min(desde + bloque.shape[0], largo)
        if hasta > desde:
            acum[desde:hasta, :] = bloque[:hasta - desde, :]

    return {
        "evento": json.loads(fila.to_json()),
        "senales": señales,
        "valores": [[None if np.isnan(v) else float(v) for v in fila_v] for fila_v in acum],
    }


# --------------------------------------------------------------------------------------
# El entregable: los Excel que produce el pipeline, pero de la rama que elijas
# --------------------------------------------------------------------------------------

#: Los tres exportables del paquete de la tesis, con el nombre base de su archivo.
EXPORTABLES = {
    "experto": ("candidatos_etapa1", "Candidatos consolidados para revisión experta"),
    "presentable": ("candidatos_etapa1_presentable", "Versión presentable, con formato y gráficos"),
    "normales": ("ventanas_normales_contraste", "Ventanas normales, para contrastar"),
}


def _mismo_valor(a, b) -> bool:
    """¿Son el mismo valor de parámetro, sin tropezarse con el tipo?

    Hace falta porque `5` y `5.0` son el mismo umbral pero no el mismo JSON. Los parámetros salen
    del backend como float, pasan por el navegador —que tiene un solo tipo numérico y los colapsa a
    entero— y vuelven así al guardarse. Comparando el texto JSON, la configuración de la tesis
    aparecía apartándose de sí misma por `percentil_baja_var=5` contra `5.0`.
    """
    # `bool` es subclase de `int`, así que va antes que la rama numérica. Comparar
    # `bool(a) is bool(b)` sería más simple pero colapsa cualquier valor no nulo a verdadero, y
    # entonces `True` y `2` pasarían por iguales. Se compara el valor, no su veracidad.
    if isinstance(a, bool) and isinstance(b, bool):
        return a is b
    if isinstance(a, bool) or isinstance(b, bool):
        return a == b
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return float(a) == float(b)
    if isinstance(a, (list, tuple)) and isinstance(b, (list, tuple)):
        return len(a) == len(b) and all(_mismo_valor(x, y) for x, y in zip(a, b))
    return a == b


def _diferencias_con_la_tesis(cadena: list[dict]) -> list[tuple[str, str, object, object]]:
    """Qué parámetros de la rama se apartan de la configuración de referencia.

    Es lo mismo que muestra la pantalla de análisis, calculado acá porque el Excel se genera en el
    servidor y tiene que poder decir de dónde salió sin depender del navegador.
    """
    dif = []
    for nodo in cadena:
        ref = ajustes.CONFIG_TESIS.get(nodo["etapa"], {})
        for k, v in (nodo["parametros"] or {}).items():
            if k in ref and not _mismo_valor(v, ref[k]):
                dif.append((nodo["etapa"], k, v, ref[k]))
    return dif


def _hoja_de_procedencia(ruta: Path, nodo: dict, cadena: list[dict], dif: list) -> None:
    """Antepone al Excel una hoja que dice de qué rama salió.

    Sin esto el taller produciría archivos indistinguibles de los oficiales de la tesis, que es
    justo lo que el ADR A7 quiere evitar: dentro de seis meses nadie podría decir si un Excel vino
    de la configuración firmada o de una prueba. La hoja va PRIMERA a propósito — es lo que se ve
    al abrir el archivo, no algo que haya que ir a buscar.
    """
    import openpyxl

    libro = openpyxl.load_workbook(ruta)
    hoja = libro.create_sheet("Procedencia", 0)
    hoja.column_dimensions["A"].width = 26
    hoja.column_dimensions["B"].width = 64

    filas = [
        ("Generado por", "Taller Etapa 1 — banco de experimentación"),
        ("", ""),
        ("¿Es la configuración de la tesis?",
         "SÍ — coincide en todos los parámetros" if not dif
         else f"NO — {len(dif)} parámetro(s) distinto(s); ver el detalle abajo"),
        ("", ""),
        ("Rama (clave)", nodo["clave"]),
        ("Commit del paquete `tesis`", nodo.get("commit_tesis") or almacen.commit_tesis()),
        ("Generado el", _ahora_iso()),
        ("", ""),
        ("ADVERTENCIA", "Las ramas del taller son exploración. El registro canónico de la tesis "
                        "—código, decisiones y resultados firmados— vive en el repositorio de la tesis."),
        ("", ""),
        ("CONFIGURACIÓN DE ESTA RAMA", ""),
    ]
    for nodo_c in cadena:
        params = ", ".join(f"{k}={v}" for k, v in sorted((nodo_c["parametros"] or {}).items()))
        filas.append((nodo_c["etapa"], params or "(sin parámetros)"))

    if dif:
        filas += [("", ""), ("SE APARTA DE LA TESIS EN", "")]
        for etapa, k, v, ref in dif:
            filas.append((f"{etapa} · {k}", f"{v}   (en la tesis: {ref})"))

    for i, (a, b) in enumerate(filas, start=1):
        hoja.cell(row=i, column=1, value=a).font = openpyxl.styles.Font(bold=True)
        hoja.cell(row=i, column=2, value=str(b))
        hoja.cell(row=i, column=2).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    libro.save(ruta)


def _ahora_iso() -> str:
    from datetime import datetime

    return datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")


def _nodo_exportable(clave: str, tipo: str) -> dict:
    """Valida que se pueda exportar esa rama con ese tipo, o corta con el error que corresponda."""
    if tipo not in EXPORTABLES:
        raise HTTPException(404, f"Tipo desconocido: {tipo}. Hay: {', '.join(EXPORTABLES)}")
    nodo = almacen.obtener(_con, clave)
    if nodo is None:
        raise HTTPException(404, "Nodo inexistente")
    if nodo["etapa"] != "eventos" or nodo["estado"] != "listo" or not nodo["tiene_resultado"]:
        raise HTTPException(409, "El entregable sale del último paso: hace falta un nodo de eventos listo")
    return nodo


def _generar_entregable(clave: str, tipo: str, carpeta: Path) -> tuple[Path, dict]:
    """Arma uno de los tres Excel del pipeline dentro de `carpeta` y devuelve (ruta, nodo).

    Está separado del endpoint porque lo usan dos caminos: la descarga suelta, que arma el archivo
    en un temporal y lo tira, y el documento de revisión, que se queda con él. Los dos tienen que
    producir exactamente lo mismo — si se duplicara la generación, con el tiempo el archivo que
    comenta el experto y el que se descarga dirían cosas distintas.

    Nada se reimplementa (ADR A1): son las mismas tres funciones del paquete de la tesis.
    """
    import dataclasses as dc

    nodo = _nodo_exportable(clave, tipo)
    salida = almacen.cargar_resultado(clave)
    cadena = almacen.cadena_hasta(_con, clave)
    dif = _diferencias_con_la_tesis(cadena)

    from tesis import export
    from tesis.config import CONFIG

    params = nodo["parametros"]
    cfg = dc.replace(
        CONFIG.detection,
        detectors=tuple(salida["scores"].columns),
        candidate_fraction=float(params["fraccion_candidatos"]),
        max_event_windows=int(params["max_ventanas_evento"]) if params["max_ventanas_evento"] else None,
    )

    base, _titulo = EXPORTABLES[tipo]
    # El nombre lleva la procedencia: quien recibe el archivo suelto tiene que poder distinguir
    # el de la configuración firmada del de una prueba sin abrirlo.
    marca = "tesis" if not dif else f"rama-{clave[:8]}"
    destino = carpeta / f"{base}_{marca}.xlsx"

    if tipo == "experto":
        export.build_expert_excel(salida["scores"], salida["meta"], destino,
                                  features=salida["filtradas"], detection_cfg=cfg,
                                  window_size=salida["tamano_ventana"])
    elif tipo == "presentable":
        export.build_presentation_excel(salida["eventos"], destino,
                                        raw_windows=salida["ventanas"], scores=salida["scores"],
                                        detection_cfg=cfg)
    else:
        normales = export.select_normal_windows(salida["scores"], salida["meta"], detection_cfg=cfg,
                                                n_examples=20, window_size=salida["tamano_ventana"])
        export.build_normal_examples_excel(normales, salida["ventanas"], destino)

    _hoja_de_procedencia(destino, nodo, cadena, dif)
    return destino, nodo


@app.get("/api/nodo/{clave}/excel/{tipo}")
def get_excel(clave: str, tipo: str):
    """Genera y devuelve uno de los Excel entregables, para la rama que se esté mirando.

    Tarda menos de un segundo, así que va en la propia petición y no por la cola de trabajos.
    """
    carpeta = Path(tempfile.mkdtemp(prefix="taller-xlsx-"))
    destino, _nodo = _generar_entregable(clave, tipo, carpeta)
    return FileResponse(
        destino, filename=destino.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        # El temporal se borra recién cuando terminó de mandarse.
        background=BackgroundTask(shutil.rmtree, carpeta, ignore_errors=True),
    )


# --------------------------------------------------------------------------------------
# Revisión experta: documentos que el experto comenta
# --------------------------------------------------------------------------------------
# El entregable de arriba es una vista: se arma, se manda y se borra. Lo de acá abajo es lo
# contrario — un artefacto que se queda, porque lo que el experto escribe encima es lo único del
# taller que no se puede volver a calcular. Ver el módulo `revision`.


class PeticionRevision(BaseModel):
    autor: str | None = None


@app.get("/api/revision/config")
def get_revision_config():
    """¿Está configurada la revisión experta, y contra qué Document Server?

    El frontend lo consulta para decidir si muestra el botón de comentar. Sin OnlyOffice el taller
    funciona igual: se pierde la pantalla de comentarios, nada más.
    """
    return {
        "habilitada": ajustes.revision_habilitada(),
        # Si está apagada, se dice por qué: es lo primero que uno quiere saber al desplegar.
        "motivo": ajustes.motivo_revision_apagada(),
        "url_publica": ajustes.url_para_el_navegador(),
        "url_del_taller": ajustes.OO_URL_DEL_TALLER,
        "jwt": bool(ajustes.OO_JWT_SECRETO),
        "tipos": {k: v[1] for k, v in EXPORTABLES.items()},
    }


@app.get("/api/revision")
def get_revisiones(nodo: str | None = None):
    """Todos los documentos de revisión, o los de una rama."""
    docs = revision.listar(nodo)
    for d in docs:
        d["editado"] = d["version"] > 0
    return {"documentos": docs}


def _con_ficha(ruta: str, doc: dict) -> str:
    """La URL de una ruta protegida, con la ficha del documento pegada."""
    return f"{ruta}?t={doc['ficha']}"


def _exigir_ficha(doc: dict, t: str | None) -> None:
    """Corta si la ficha no es la del documento.

    Protege las dos rutas que tienen que quedar abiertas en el proxy de adelante porque las usa el
    Document Server, que no puede autenticarse. La del callback es la que importa: sobrescribe el
    documento del experto.
    """
    if not revision.ficha_valida(doc, t):
        raise HTTPException(403, "Ficha inválida o ausente para ese documento")


def _doc_publico(doc: dict) -> dict:
    """El documento como lo ve el frontend, con su historial."""
    salida = dict(doc)
    salida["editado"] = doc["version"] > 0
    salida["versiones"] = revision.versiones(doc["id"])
    return salida


@app.post("/api/nodo/{clave}/revision/{tipo}")
def post_revision(clave: str, tipo: str, pet: PeticionRevision | None = None):
    """Abre el documento de revisión de esa rama, creándolo la primera vez.

    **Es idempotente a propósito.** Si ya existe, devuelve el que hay sin tocarlo: volver a entrar a
    comentar no puede regenerar el archivo encima de lo que el experto ya escribió. Por eso el
    entregable normal y este documento son cosas distintas aunque salgan de la misma función.
    """
    doc_id = revision.id_documento(clave, tipo)
    doc = revision.obtener(doc_id)

    # Si ya lo comentó, se abre lo que hay y no se toca. Esa es la regla.
    if doc is not None and doc["version"] > 0:
        return _doc_publico(doc)

    # Pero si está en v0 nadie escribió nada todavía: es una copia del entregable y nada más. En ese
    # caso se regenera, para que las correcciones al pipeline lleguen al experto en vez de dejarlo
    # mirando una versión vieja. (Pasó con los gráficos: se arregló la posición de los ejes en
    # `tesis.export` y los documentos ya creados habrían seguido mostrando el error para siempre.)
    #
    # No contradice la regla de «no regenerar encima»: acá no hay nada encima que perder.
    if doc is not None:
        revision.olvidar(doc_id)

    nodo = _nodo_exportable(clave, tipo)
    carpeta = Path(tempfile.mkdtemp(prefix="taller-revision-"))
    try:
        origen, _ = _generar_entregable(clave, tipo, carpeta)
        doc = revision.crear(
            nodo=clave, tipo=tipo, nombre=origen.name, origen=origen,
            commit_tesis=nodo.get("commit_tesis") or almacen.commit_tesis(),
        )
    finally:
        shutil.rmtree(carpeta, ignore_errors=True)
    return _doc_publico(doc)


@app.get("/api/revision/{doc_id}")
def get_revision(doc_id: str):
    doc = revision.obtener(doc_id)
    if doc is None:
        raise HTTPException(404, "No hay documento de revisión con ese id")
    return _doc_publico(doc)


@app.get("/api/revision/{doc_id}/editor")
def get_revision_editor(doc_id: str, autor: str = "Experto"):
    """La configuración que el navegador le pasa al editor de OnlyOffice.

    Se arma en el backend y no en el JavaScript porque acá están las direcciones: el navegador no
    tiene por qué saber con qué nombre el Document Server ve al taller, que además es distinto en
    desarrollo y en el servidor.
    """
    doc = revision.obtener(doc_id)
    if doc is None:
        raise HTTPException(404, "No hay documento de revisión con ese id")
    if not ajustes.revision_habilitada():
        raise HTTPException(503, "La revisión experta no está configurada (falta OO_URL_PUBLICA)")

    autor = (autor or "Experto").strip()[:60] or "Experto"
    base = ajustes.OO_URL_DEL_TALLER
    config = {
        "document": {
            "fileType": "xlsx",
            # La `key` lleva la versión adentro: si no cambiara al guardar, el Document Server
            # seguiría sirviendo su copia en caché y el experto abriría una versión vieja.
            "key": revision.clave_oo(doc),
            "title": doc["nombre"],
            "url": base + _con_ficha(f"/api/revision/{doc_id}/archivo", doc),
        },
        "documentType": "cell",
        "editorConfig": {
            "callbackUrl": base + _con_ficha(f"/api/revision/{doc_id}/callback", doc),
            "lang": ajustes.OO_IDIOMA,
            "user": {"id": autor, "name": autor},
            "customization": {"autosave": True, "forcesave": True, "comments": True},
            # OnlyOffice arranca solo once plugins: ai, ocr, photoeditor, speech, translator,
            # youtube, zotero, mendeley, thesaurus, highlightcode y speechrecognition. Entre todos
            # son 1890 archivos (43 MB), y el de IA solo son 786 —un script por cada proveedor de
            # modelos, más un vocabulario de tokenización—. Nada de eso sirve para anotar una
            # planilla de eventos candidatos.
            #
            # Con la lista vacía no se carga ninguno. Lo que se gana es la primera apertura: es la
            # que el experto va a sufrir, porque cada reinicio del Document Server cambia su huella
            # de compilación y vuelve a enfriar la caché del navegador.
            "plugins": {"autostart": [], "pluginsData": []},
        },
    }
    return {"url_publica": ajustes.url_para_el_navegador(), "config": config}


@app.get("/api/revision/{doc_id}/archivo")
def get_revision_archivo(doc_id: str, t: str | None = None):
    """Sirve el Excel vigente. Lo piden dos: el Document Server al abrir, y quien lo descargue."""
    doc = revision.obtener(doc_id)
    if doc is None:
        raise HTTPException(404, "No hay documento de revisión con ese id")
    _exigir_ficha(doc, t)
    ruta = revision.ruta_actual(doc_id)
    if not ruta.exists():
        raise HTTPException(410, "El archivo del documento no está en disco")
    return FileResponse(
        ruta, filename=doc["nombre"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/revision/{doc_id}/version/{numero}")
def get_revision_version(doc_id: str, numero: int, t: str | None = None):
    """Baja una versión concreta. La 0 es lo que salió del pipeline, antes de que nadie lo tocara."""
    doc = revision.obtener(doc_id)
    if doc is None:
        raise HTTPException(404, "No hay documento de revisión con ese id")
    _exigir_ficha(doc, t)
    ruta = revision.ruta_version(doc_id, numero)
    if not ruta.exists():
        raise HTTPException(404, f"No existe la versión {numero}")
    return FileResponse(
        ruta, filename=f"v{numero:04d}_{doc['nombre']}",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _bajar_del_document_server(url: str) -> bytes:
    """Baja el archivo guardado desde el Document Server.

    La URL viene del editor y apunta a sí mismo con un nombre que a veces solo él resuelve. Si está
    configurada `OO_URL_INTERNA`, se le reemplaza el origen por esa — que es con la que el taller
    sabe llegar.
    """
    import urllib.parse
    import urllib.request

    if ajustes.OO_URL_INTERNA:
        partes = urllib.parse.urlsplit(url)
        base = urllib.parse.urlsplit(ajustes.OO_URL_INTERNA)
        url = urllib.parse.urlunsplit(
            (base.scheme or partes.scheme, base.netloc, partes.path, partes.query, partes.fragment)
        )
    with urllib.request.urlopen(url, timeout=60) as r:
        return r.read()


def _reindexar(doc_id: str) -> dict:
    """Vuelve a leer el archivo vigente y deja el índice al día."""
    resumen = revision.resumen_de_comentarios(
        revision.ruta_actual(doc_id),
        revision.columnas_originales(revision.ruta_version(doc_id, 0)),
    )
    revision.guardar_indice(doc_id, resumen)
    return resumen


@app.post("/api/revision/{doc_id}/callback")
async def post_revision_callback(doc_id: str, request: Request, t: str | None = None):
    """Lo que el Document Server llama cuando pasa algo con el documento.

    Del protocolo de OnlyOffice interesan dos estados: 2 (listo para guardar, ya cerraron todos) y
    6 (guardado forzado, con la sesión todavía abierta). En los dos casos el editor **no manda el
    archivo**: manda una URL suya para que lo vayamos a buscar.

    Hay que contestar `{"error": 0}` siempre que se haya manejado bien, incluso en los estados que
    se ignoran; si no, el Document Server reintenta y termina dando el documento por fallado.
    """
    try:
        cuerpo = await request.json()
    except Exception:
        raise HTTPException(400, "El cuerpo del callback no es JSON")

    doc = revision.obtener(doc_id)
    if doc is None:
        raise HTTPException(404, "No hay documento de revisión con ese id")
    _exigir_ficha(doc, t)

    estado = int(cuerpo.get("status", 0))
    if estado not in (2, 6):
        # 1 = alguien está editando · 4 = cerraron sin cambios · 3 y 7 = error del editor.
        # Ninguno mueve el archivo, pero igual se contesta OK para que no reintente.
        return {"error": 0}

    url = cuerpo.get("url")
    if not url:
        raise HTTPException(400, "El callback pide guardar pero no manda la URL del archivo")

    try:
        contenido = _bajar_del_document_server(url)
    except Exception as e:
        # Se devuelve error a propósito para que el Document Server reintente: prefiero que insista
        # antes que dar por guardado algo que no bajó.
        raise HTTPException(502, f"No se pudo bajar el archivo guardado: {type(e).__name__}: {e}")

    usuarios = cuerpo.get("users") or []
    autor = str(usuarios[0])[:60] if usuarios else None
    revision.registrar_guardado(doc_id, contenido, autor)

    # El índice es derivado: si falla, el guardado igual valió, y se puede volver a generar desde
    # las versiones que quedaron en disco. No se pierde nada.
    try:
        _reindexar(doc_id)
    except Exception:
        pass

    return {"error": 0}


@app.get("/api/revision/{doc_id}/comentarios")
def get_revision_comentarios(doc_id: str, releer: bool = False):
    """Lo que el experto escribió, extraído del archivo.

    **Es un índice, no la fuente de verdad**: el archivo manda. Sirve para poder buscar sin abrir
    cuarenta planillas. Con `releer=1` se vuelve a pasar sobre el archivo vigente, que es lo que hay
    que hacer si se mejora la extracción.
    """
    doc = revision.obtener(doc_id)
    if doc is None:
        raise HTTPException(404, "No hay documento de revisión con ese id")

    resumen = None if releer else revision.leer_indice(doc_id)
    if resumen is None:
        if not revision.ruta_actual(doc_id).exists():
            raise HTTPException(410, "El archivo del documento no está en disco")
        resumen = _reindexar(doc_id)
    return {"documento": doc_id, "version": doc["version"], **resumen}

# --------------------------------------------------------------------------------------
# Web estática
# --------------------------------------------------------------------------------------

@app.middleware("http")
async def revalidar_el_frontend(request, call_next):
    """Obliga al navegador a preguntar siempre si el HTML/CSS/JS cambió.

    Sin esto, las respuestas del frontend salen con `etag` y `last-modified` pero **sin**
    `Cache-Control`. Ante esa combinación el navegador aplica «frescura heurística»: decide por su
    cuenta durante cuánto tiempo el archivo sigue vigente y lo sirve de su caché sin preguntar
    nada. El resultado es que editar `web/` y recargar dejaba media aplicación vieja — típicamente
    una pantalla con la versión nueva y la otra con la anterior, que es difícil de diagnosticar
    porque el servidor sí está entregando el archivo correcto.

    Eso choca de frente con que el frontend se monte en el contenedor justamente para poder
    editarlo sin reconstruir la imagen.

    `no-cache` no prohíbe guardar: obliga a revalidar. Para `/estatico/*` eso sale gratis, porque
    StaticFiles contesta 304 vacío ante el `etag`. Las dos rutas de HTML reenvían el cuerpo entero
    (unos 30 KB entre las dos, contra un servidor que corre en la misma máquina): se prefiere eso
    antes que volver a tener pantallas desincronizadas.

    La API queda afuera: sus respuestas son dinámicas y ya nadie las cachea.
    """
    respuesta = await call_next(request)
    if not request.url.path.startswith("/api/"):
        respuesta.headers["Cache-Control"] = "no-cache"
    return respuesta


@app.get("/")
def raiz():
    return FileResponse(WEB / "index.html")


@app.get("/taller")
def taller():
    return FileResponse(WEB / "taller.html")


@app.get("/revision")
def revision_html():
    """La pantalla donde el experto comenta. Se abre con `?doc=<id>`."""
    return FileResponse(WEB / "revision.html")


app.mount("/estatico", StaticFiles(directory=WEB), name="estatico")


# El proxy del Document Server va ÚLTIMO, y no es un detalle: su ruta es un comodín `/{ruta:path}`,
# así que registrado antes se tragaría todo lo del taller. Starlette resuelve por orden de registro.
if ajustes.OO_PROXY:
    app.include_router(proxy_oo.router)


@app.on_event("shutdown")
async def _cerrar_proxy():
    await proxy_oo.cerrar()
