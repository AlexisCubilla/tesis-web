"""Revisión experta: el Excel que el experto comenta, y todo lo que escribió en él.

Este módulo guarda lo único del taller que **no se puede recalcular**. Cada rama, cada número y cada
gráfico salen del pipeline: si se pierden, se vuelven a correr. El juicio del experto no. Es trabajo
humano sobre telemetría que nadie etiquetó nunca, y es la materia prima de la Etapa 2.

De ahí las tres reglas que sigue todo lo de acá abajo:

1. **El documento se crea una vez y no se regenera encima.** El entregable normal
   (`/api/nodo/{clave}/excel/{tipo}`) se arma en un temporal y se borra al terminar de mandarse: es
   una vista. El documento de revisión es lo contrario — es un artefacto que se queda. Volver a
   correr la rama NO lo toca; si hiciera falta uno nuevo, se crea aparte.

2. **Cada guardado deja una versión.** El Document Server manda el archivo entero en cada guardado,
   así que guardarlos todos cuesta unos KB y elimina la categoría entera de «se perdió»: un borrado
   accidental, una sesión caída o un cierre a destiempo se arreglan volviendo una versión atrás.

3. **El archivo manda.** Lo que el experto escribió es el archivo, no una interpretación nuestra.
   `resumen_de_comentarios()` lo lee para poder buscar sin abrir cuarenta planillas, pero eso es un
   índice derivado: si la extracción está mal, se corrige y se vuelve a pasar sobre las versiones
   guardadas, sin perder nada.

El experto trabaja con libertad total sobre la planilla — columnas nuevas, hojas nuevas, colores,
comentarios de celda. No hay formulario que lo encajone, porque la taxonomía de la Etapa 2 es
justamente lo que él tiene que producir; ofrecerle una lista cerrada sería adelantarle la respuesta.
"""

from __future__ import annotations

import hashlib
import json
import shutil
import sqlite3
import threading
from datetime import datetime, timezone
from pathlib import Path

from . import ajustes

#: Carpeta raíz de los documentos de revisión. Va dentro de DIR_DATOS como todo el estado local,
#: pero a diferencia del caché NO se puede borrar para «empezar de cero»: acá está lo irreemplazable.
DIR_REVISION: Path = ajustes.DIR_DATOS / "revision"

#: Base propia, deliberadamente separada de `ejecuciones.sqlite`.
#:
#: El árbol de ramas es desechable: `mise run limpiar` lo borra entero, y está bien, porque se
#: recalcula corriendo el pipeline de nuevo. Si el registro de los documentos viviera en ese mismo
#: archivo, esa tarea se llevaría puesto el trabajo del experto —que no se recalcula— y dejaría los
#: .xlsx en disco sin nada que diga de qué rama salieron ni qué versión es cuál.
#:
#: Dos ciclos de vida distintos, dos bases distintas.
BASE: Path = DIR_REVISION / "revision.sqlite"

_lock = threading.Lock()
_con: sqlite3.Connection | None = None

ESQUEMA = """
CREATE TABLE IF NOT EXISTS documentos (
    id            TEXT PRIMARY KEY,
    ficha         TEXT NOT NULL DEFAULT '',
    nodo          TEXT NOT NULL,
    tipo          TEXT NOT NULL,
    nombre        TEXT NOT NULL,
    version       INTEGER NOT NULL DEFAULT 0,
    creado_en     TEXT NOT NULL,
    guardado_en   TEXT,
    ultimo_autor  TEXT,
    commit_tesis  TEXT,
    rama_borrada  INTEGER NOT NULL DEFAULT 0
);
CREATE INDEX IF NOT EXISTS idx_documentos_nodo ON documentos(nodo);

CREATE TABLE IF NOT EXISTS versiones (
    documento   TEXT NOT NULL,
    version     INTEGER NOT NULL,
    archivo     TEXT NOT NULL,
    autor       TEXT,
    guardado_en TEXT NOT NULL,
    bytes       INTEGER NOT NULL,
    PRIMARY KEY (documento, version)
);
"""


def _ahora() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def preparar() -> sqlite3.Connection:
    """Abre la base de revisión, crea sus tablas y pone al día las que ya existan."""
    global _con
    DIR_REVISION.mkdir(parents=True, exist_ok=True)
    _con = sqlite3.connect(BASE, check_same_thread=False)
    _con.row_factory = sqlite3.Row
    _con.executescript(ESQUEMA)

    # Bases creadas antes de que existiera la ficha: se agrega la columna y se le da una a cada
    # documento. Sin esto, los documentos viejos quedarían sin secreto y no se podrían abrir.
    columnas = {f["name"] for f in _con.execute("PRAGMA table_info(documentos)")}
    if "ficha" not in columnas:
        _con.execute("ALTER TABLE documentos ADD COLUMN ficha TEXT NOT NULL DEFAULT ''")
    faltantes = [f["id"] for f in _con.execute("SELECT id FROM documentos WHERE ficha = ''")]
    for doc_id in faltantes:
        _con.execute("UPDATE documentos SET ficha=? WHERE id=?", (nueva_ficha(), doc_id))
    if faltantes:
        _con.commit()
    return _con


def nueva_ficha() -> str:
    """Un secreto por documento. 32 bytes: no se adivina."""
    import secrets

    return secrets.token_urlsafe(32)


def ficha_valida(doc: dict, entregada: str | None) -> bool:
    """¿La ficha entregada es la del documento? Comparación en tiempo constante."""
    import secrets

    esperada = doc.get("ficha") or ""
    if not esperada or not entregada:
        return False
    return secrets.compare_digest(esperada, entregada)


def conexion() -> sqlite3.Connection:
    """La conexión a la base de revisión, abriéndola si hiciera falta."""
    return _con if _con is not None else preparar()


# --------------------------------------------------------------------------------------
# Identidad
# --------------------------------------------------------------------------------------

def id_documento(nodo: str, tipo: str) -> str:
    """Un documento por (rama, tipo de entregable). Estable: reabrir lleva al mismo archivo."""
    return f"{nodo}-{tipo}"


def clave_oo(doc: dict) -> str:
    """La `key` que ve el Document Server.

    Tiene que cambiar cada vez que cambia el contenido, o el servidor sirve su copia en caché y el
    experto abre una versión vieja sin enterarse. Por eso lleva el número de versión adentro.

    OnlyOffice acepta hasta 128 caracteres y solo `[0-9a-zA-Z._=-]`, así que el hash evita que un
    nombre de rama con cualquier cosa adentro rompa el editor.
    """
    crudo = f"{doc['id']}|{doc['version']}"
    return hashlib.sha1(crudo.encode()).hexdigest() + f"-v{doc['version']}"


def _carpeta(doc_id: str) -> Path:
    return DIR_REVISION / doc_id


def ruta_actual(doc_id: str) -> Path:
    """El archivo vigente: el que se le sirve al Document Server y el que se descarga."""
    return _carpeta(doc_id) / "actual.xlsx"


def ruta_version(doc_id: str, version: int) -> Path:
    return _carpeta(doc_id) / "versiones" / f"{version:04d}.xlsx"


# --------------------------------------------------------------------------------------
# Ciclo de vida
# --------------------------------------------------------------------------------------

def obtener(doc_id: str) -> dict | None:
    fila = conexion().execute("SELECT * FROM documentos WHERE id=?", (doc_id,)).fetchone()
    return dict(fila) if fila else None


def crear(
    *,
    nodo: str,
    tipo: str,
    nombre: str,
    origen: Path,
    commit_tesis: str | None,
) -> dict:
    """Registra un documento nuevo copiando el Excel recién generado, y lo guarda como versión 0.

    La versión 0 es «lo que salió del pipeline, antes de que nadie lo tocara». Sirve para poder
    contestar más adelante qué agregó el experto y qué ya estaba.
    """
    doc_id = id_documento(nodo, tipo)
    carpeta = _carpeta(doc_id)
    (carpeta / "versiones").mkdir(parents=True, exist_ok=True)

    shutil.copy2(origen, ruta_actual(doc_id))
    shutil.copy2(origen, ruta_version(doc_id, 0))

    ahora = _ahora()
    con = conexion()
    with _lock:
        con.execute(
            "INSERT INTO documentos (id, ficha, nodo, tipo, nombre, version, creado_en, commit_tesis) "
            "VALUES (?,?,?,?,?,0,?,?)",
            (doc_id, nueva_ficha(), nodo, tipo, nombre, ahora, commit_tesis),
        )
        con.execute(
            "INSERT INTO versiones (documento, version, archivo, autor, guardado_en, bytes) "
            "VALUES (?,0,?,?,?,?)",
            (doc_id, str(ruta_version(doc_id, 0)), None, ahora,
             ruta_version(doc_id, 0).stat().st_size),
        )
        con.commit()
    return obtener(doc_id)


def registrar_guardado(doc_id: str, contenido: bytes, autor: str | None) -> dict:
    """Guarda lo que mandó el Document Server como una versión nueva y la deja como vigente.

    Nunca se pisa una versión anterior: el número sube y el archivo viejo se queda donde está.
    """
    doc = obtener(doc_id)
    if doc is None:
        raise KeyError(doc_id)

    version = int(doc["version"]) + 1
    destino = ruta_version(doc_id, version)
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_bytes(contenido)
    # El vigente se escribe después de que la versión ya está en disco: si algo falla en el medio,
    # se pierde el guardado nuevo, nunca el anterior.
    shutil.copy2(destino, ruta_actual(doc_id))

    ahora = _ahora()
    con = conexion()
    with _lock:
        con.execute(
            "INSERT INTO versiones (documento, version, archivo, autor, guardado_en, bytes) "
            "VALUES (?,?,?,?,?,?)",
            (doc_id, version, str(destino), autor, ahora, len(contenido)),
        )
        con.execute(
            "UPDATE documentos SET version=?, guardado_en=?, ultimo_autor=? WHERE id=?",
            (version, ahora, autor, doc_id),
        )
        con.commit()
    return obtener(doc_id)


def olvidar(doc_id: str) -> None:
    """Borra un documento **sin ediciones** para poder volver a generarlo.

    Solo se llama sobre documentos en versión 0, que son una copia del entregable y nada más. Si
    tuviera aunque sea un guardado del experto, esto se llevaría trabajo irrepetible — por eso la
    comprobación vive en quien llama y acá se deja dicho: **no llamar sobre un documento editado**.
    """
    doc = obtener(doc_id)
    if doc is None:
        return
    if doc["version"] > 0:
        raise ValueError(f"{doc_id} tiene ediciones del experto (v{doc['version']}): no se borra")

    shutil.rmtree(_carpeta(doc_id), ignore_errors=True)
    con = conexion()
    with _lock:
        con.execute("DELETE FROM versiones WHERE documento=?", (doc_id,))
        con.execute("DELETE FROM documentos WHERE id=?", (doc_id,))
        con.commit()


def versiones(doc_id: str) -> list[dict]:
    filas = conexion().execute(
        "SELECT version, autor, guardado_en, bytes FROM versiones WHERE documento=? "
        "ORDER BY version DESC",
        (doc_id,),
    ).fetchall()
    return [dict(f) for f in filas]


def listar(nodo: str | None = None) -> list[dict]:
    con = conexion()
    if nodo is None:
        filas = con.execute("SELECT * FROM documentos ORDER BY creado_en DESC").fetchall()
    else:
        filas = con.execute(
            "SELECT * FROM documentos WHERE nodo=? ORDER BY creado_en DESC", (nodo,)
        ).fetchall()
    return [dict(f) for f in filas]


# --------------------------------------------------------------------------------------
# Protección al borrar ramas
# --------------------------------------------------------------------------------------

def con_trabajo_del_experto(claves: list[str]) -> list[dict]:
    """De esas ramas, cuáles tienen un documento que el experto ya editó (versión > 0).

    Es la pregunta que hay que hacer antes de borrar: el árbol de ramas se puede reconstruir corriendo
    el pipeline de nuevo, pero lo que escribió el experto no. Un documento sin editar (versión 0) no
    cuenta: es una copia del entregable y se vuelve a generar solo.
    """
    if not claves:
        return []
    marcas = ",".join("?" * len(claves))
    filas = conexion().execute(
        f"SELECT * FROM documentos WHERE nodo IN ({marcas}) AND version > 0 "
        "ORDER BY guardado_en DESC",
        claves,
    ).fetchall()
    return [dict(f) for f in filas]


def desvincular(claves: list[str]) -> int:
    """Marca los documentos de esas ramas como huérfanos, en vez de borrarlos con la rama.

    Se usa cuando alguien borra igual una rama comentada: la rama se va, el archivo del experto se
    queda. Un documento huérfano ya no se puede abrir contra su rama —no hay contra qué compararlo—
    pero se sigue pudiendo descargar, y la hoja de Procedencia de adentro dice de dónde salió.
    """
    if not claves:
        return 0
    marcas = ",".join("?" * len(claves))
    con = conexion()
    with _lock:
        cur = con.execute(
            f"UPDATE documentos SET rama_borrada=1 WHERE nodo IN ({marcas})", claves
        )
        con.commit()
    return cur.rowcount


# --------------------------------------------------------------------------------------
# Índice derivado de lo que escribió el experto
# --------------------------------------------------------------------------------------

def _cabecera(hoja) -> list[str]:
    """Los nombres de columna de la primera fila."""
    filas = hoja.iter_rows(max_row=1)
    try:
        primera = next(filas)
    except StopIteration:  # hoja vacía
        return []
    return [str(c.value) if c.value is not None else "" for c in primera]


def columnas_originales(ruta_v0: Path) -> dict[str, list[str]]:
    """Qué columnas y hojas traía el entregable recién salido del pipeline.

    Se compara contra esto para saber qué agregó el experto. Sale de la versión 0 y no de una lista
    escrita a mano acá, porque las columnas del entregable las decide `tesis.export`: si allá se
    agrega una, no habría forma de enterarse y aparecería como si la hubiera escrito el experto.
    """
    import openpyxl

    if not ruta_v0.exists():
        return {}
    libro = openpyxl.load_workbook(ruta_v0, read_only=True)
    try:
        return {hoja.title: _cabecera(hoja) for hoja in libro.worksheets}
    finally:
        libro.close()


def resumen_de_comentarios(ruta: Path, columnas_originales: dict[str, list[str]] | None = None) -> dict:
    """Lee el archivo y junta todo lo que parezca escrito por una persona.

    **Esto es un índice, no la fuente de verdad.** Sirve para buscar sin abrir cuarenta planillas y
    para que la Etapa 3 tenga algo que leer que no sea arqueología de XML. El archivo manda: si acá
    se extrae mal, se corrige la función y se vuelve a pasar sobre las versiones guardadas.

    Recoge tres cosas, porque no se le impone al experto dónde escribir:
      · comentarios de celda (los globitos del editor),
      · columnas que no estaban en el entregable original,
      · hojas nuevas que agregó.
    """
    import openpyxl

    libro = openpyxl.load_workbook(ruta)
    originales = columnas_originales or {}

    comentarios: list[dict] = []
    columnas_nuevas: dict[str, list[str]] = {}
    hojas_nuevas: list[str] = []

    for hoja in libro.worksheets:
        conocidas = originales.get(hoja.title)
        if conocidas is None and originales:
            hojas_nuevas.append(hoja.title)

        for fila in hoja.iter_rows():
            for celda in fila:
                if celda.comment is not None and str(celda.comment.text).strip():
                    comentarios.append({
                        "hoja": hoja.title,
                        "celda": celda.coordinate,
                        "fila": celda.row,
                        "autor": (celda.comment.author or "").strip() or None,
                        "texto": str(celda.comment.text).strip(),
                    })

        if conocidas:
            cabecera = _cabecera(hoja)
            agregadas = [c for c in cabecera if c and c not in conocidas]
            if agregadas:
                columnas_nuevas[hoja.title] = agregadas

    texto_libre = _texto_de_columnas_nuevas(libro, columnas_nuevas)

    return {
        "comentarios_de_celda": comentarios,
        "columnas_agregadas": columnas_nuevas,
        "hojas_agregadas": hojas_nuevas,
        "texto_en_columnas_nuevas": texto_libre,
        "total": len(comentarios) + len(texto_libre),
    }


def _texto_de_columnas_nuevas(libro, columnas_nuevas: dict[str, list[str]]) -> list[dict]:
    """Lo que el experto escribió en columnas que él mismo agregó."""
    salida: list[dict] = []
    for titulo, nombres in columnas_nuevas.items():
        hoja = libro[titulo]
        cabecera = _cabecera(hoja)
        indices = {n: cabecera.index(n) for n in nombres if n in cabecera}
        for fila in hoja.iter_rows(min_row=2):
            for nombre, i in indices.items():
                if i >= len(fila):
                    continue
                valor = fila[i].value
                if valor is None or not str(valor).strip():
                    continue
                salida.append({
                    "hoja": titulo,
                    "fila": fila[0].row,
                    "columna": nombre,
                    "texto": str(valor).strip(),
                })
    return salida


def guardar_indice(doc_id: str, resumen: dict) -> None:
    """Deja el índice al lado del archivo. Va en disco y no en una tabla porque es derivado:
    se puede borrar y volver a generar desde las versiones guardadas."""
    destino = _carpeta(doc_id) / "indice.json"
    destino.write_text(json.dumps(resumen, ensure_ascii=False, indent=2), encoding="utf-8")


def leer_indice(doc_id: str) -> dict | None:
    ruta = _carpeta(doc_id) / "indice.json"
    if not ruta.exists():
        return None
    return json.loads(ruta.read_text(encoding="utf-8"))
